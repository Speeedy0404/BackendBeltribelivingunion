import os
import sys
from django.conf import settings
from openpyxl import load_workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from ..models import Farms, PK, PKYoungAnimals, PKBull, Report, JsonFarmsData
from ..serializers import CowParameterForecastingSerializer, BullParameterForecastingSerializer

REPORT_DIR = os.path.join(settings.BASE_DIR, 'reports')
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))

from fields import MAPPING


def calculate_weighted_average(bull_data, index_type, number):
    total_ebv_rel = {}
    total_ebv = {}
    total_rel = {}

    for bull in bull_data:
        index = bull[index_type]
        if index is None:
            pass
        else:
            for parameter, value in index.items():
                if parameter.startswith('ebv_'):
                    rel_parameter = parameter.replace('ebv_', 'rel_')
                    rel_value = index[rel_parameter]

                    if value is None or rel_value is None:
                        continue

                    ebv_rel_value = value * (rel_value / 100)
                    ebv_value = value

                    if parameter not in total_ebv_rel:
                        total_ebv_rel[parameter] = 0
                        total_ebv[parameter] = 0
                        total_rel[parameter] = 0

                    total_ebv_rel[parameter] += ebv_rel_value
                    total_ebv[parameter] += ebv_value
                    total_rel[parameter] += 1

    if len(total_ebv_rel) == 0 or len(total_rel) == 0:
        return {'param': 0}, {'param': 0}
    else:
        average_values = {parameter: total_ebv_rel[parameter] / total_rel[parameter] for parameter in total_ebv_rel}
        average_bull = {parameter.split('_')[-1]: (total_ebv[parameter] / total_rel[parameter]) * number for parameter
                        in
                        total_ebv}
    return average_values, average_bull


def calculate_weighted_average_with_bulls(cow_data, averages, index_type):
    cows = []

    for cow in cow_data:
        count = -1
        cow_averages = {}
        for index in index_type:
            count += 1
            cow_index = cow.get(index)
            bull_average = averages[count]
            if cow_index is None:
                cow_averages[index] = None
            else:
                current_param = {}
                for parameter, value in cow_index.items():
                    if parameter.startswith('ebv_'):
                        rel_parameter = parameter.replace('ebv_', 'rel_')
                        rel_value = cow_index.get(rel_parameter)
                        if value is None or rel_value is None:
                            current_param[parameter] = None
                        else:
                            average_need = bull_average.get(parameter, 0)
                            ebv_rel_value = ((value * (rel_value / 100)) + average_need) / 2
                            current_param[parameter] = ebv_rel_value
                cow_averages[index] = current_param
        cows.append(cow_averages)
    return cows


def calculate_average(params_list, data, param_key):
    total = 0
    count = 0
    for cow in params_list:
        value = cow.get(data, {})
        if value is None:
            pass
        else:
            value = value.get(param_key, None)

        if value is not None:
            total += value
            count += 1
    return total / count if count > 0 else 0


def count_valid_conformationindex(params_list):
    count = 0
    for cow in params_list:
        conformationindex = cow.get('conformationindex', None)
        if conformationindex is not None:
            count += 1
    return count


def set_predict(current, forecasting, avg_dict):
    for element in current:
        key = element['param']
        element['predict'] = forecasting.get(key, 0)
        element['bull_superiority'] = avg_dict.get(key, 0) - element['avg']
    return current


def set_predict_null(data):
    for element in data:
        if element['param'] != 0:
            element['predict'] = 0
            element['bull_superiority'] = 0
    return data


def mapping_label(group_list):
    for item in group_list:
        param_name = item['param']
        item['param'] = MAPPING[param_name]
    return group_list


def get_weighted_avg_bull(avg_values, count_cows):
    bull_averages = {}
    for element in avg_values:
        for sub_element in element:
            for key, value in sub_element.items():
                if key not in bull_averages:
                    bull_averages[key] = 0
                bull_averages[key] += value
    for key, value in bull_averages.items():
        bull_averages[key] = value / count_cows
    return bull_averages


class ParameterForecastingView(APIView):
    def post(self, request):
        try:
            cows_param = []
            cow_numbers = []
            bull_numbers = []
            avg_values = []
            count_cows = 0
            farm = Farms.objects.get(korg=self.request.headers.get('Kodrn'))
            paths = Report.objects.filter(title__icontains=farm.norg).values_list('path', flat=True)
            aggregated_data = farm.jsonfarmsdata.aggregated_data['aggregated_data']

            if len(paths) == 0:
                forecasting_1 = set_predict_null(aggregated_data['forecasting_section_one'])
                forecasting_2 = set_predict_null(aggregated_data['forecasting_section_two'])
                forecasting_3 = set_predict_null(aggregated_data['forecasting_section_three'])
                forecasting_4 = set_predict_null(aggregated_data['forecasting_section_four'])
                farm.jsonfarmsdata.save()
                forecasting_1 = mapping_label(forecasting_1)
                forecasting_2 = mapping_label(forecasting_2)
                forecasting_3 = mapping_label(forecasting_3)
                forecasting_4 = mapping_label(forecasting_4)
                result_data = {
                    'forecasting_data_one': forecasting_1,
                    'forecasting_data_two': forecasting_2,
                    'forecasting_data_thee': forecasting_3,
                    'forecasting_data_four': forecasting_4,
                }
                return Response(result_data, status=status.HTTP_200_OK)

            new_paths = [path + ".xlsx" for path in paths]
            directory_path = new_paths[0].split('__')[0][:-16]

            reports = []
            for new_path in new_paths:
                xlsx_path = os.path.join(REPORT_DIR, directory_path, new_path)
                if os.path.exists(xlsx_path):
                    workbook = load_workbook(xlsx_path)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=5, min_col=1, max_col=1, values_only=True):
                        if row[0]:
                            cow_numbers.append(row[0])
                    for row in sheet.iter_rows(min_row=5, min_col=7, max_col=7, values_only=True):
                        if row[0]:
                            bull_numbers.append(row[0])
                reports.append([cow_numbers, bull_numbers])
                cow_numbers = []
                bull_numbers = []

            for report in reports:
                cows = PK.objects.filter(uniq_key__in=report[0]).values_list('uniq_key', flat=True)

                if len(cows) == len(report[0]):
                    queryset = PK.objects.filter(
                        uniq_key__in=cows
                    ).select_related(
                        'milkproductionindex', 'conformationindex', 'reproductionindex',
                        'somaticcellindex',
                    ).order_by('id')

                    number = queryset.count()
                    count_cows += number
                    serializer = CowParameterForecastingSerializer(queryset, many=True)
                    cow_data = serializer.data

                    queryset = PKBull.objects.filter(
                        uniq_key__in=report[1]
                    ).select_related(
                        'milkproductionindexbull', 'conformationindexbull', 'reproductionindexbull',
                        'somaticcellindexbull',
                    ).order_by('id')

                    serializer = BullParameterForecastingSerializer(queryset, many=True)
                    bull_data = serializer.data

                    average_conformation, avg_conf = calculate_weighted_average(bull_data, 'conformationindexbull',
                                                                                number)
                    average_milk, avg_milk = calculate_weighted_average(bull_data, 'milkproductionindexbull', number)
                    average_reproduction, avg_reprod = calculate_weighted_average(bull_data, 'reproductionindexbull',
                                                                                  number)
                    average_somaticcell, avg_somatic = calculate_weighted_average(bull_data, 'somaticcellindexbull',
                                                                                  number)

                    avg_values.append([avg_conf, avg_milk, avg_reprod, avg_somatic])
                    cows = calculate_weighted_average_with_bulls(cow_data,
                                                                 [average_conformation, average_milk,
                                                                  average_reproduction,
                                                                  average_somaticcell],
                                                                 ['conformationindex', 'milkproductionindex',
                                                                  'reproductionindex', 'somaticcellindex'])

                    cows_param.extend(cows)
                else:
                    pass
            avg_dict = get_weighted_avg_bull(avg_values, count_cows)

            forecasting = {
                'tip': calculate_average(cows_param, 'conformationindex', 'ebv_tip'),
                'kt': calculate_average(cows_param, 'conformationindex', 'ebv_kt'),
                'rost': calculate_average(cows_param, 'conformationindex', 'ebv_rost'),
                'gt': calculate_average(cows_param, 'conformationindex', 'ebv_gt'),
                'pz': calculate_average(cows_param, 'conformationindex', 'ebv_pz'),
                'shz': calculate_average(cows_param, 'conformationindex', 'ebv_shz'),
                'pzkb': calculate_average(cows_param, 'conformationindex', 'ebv_pzkb'),
                'pzkz': calculate_average(cows_param, 'conformationindex', 'ebv_pzkz'),
                'sust': calculate_average(cows_param, 'conformationindex', 'ebv_sust'),
                'pzkop': calculate_average(cows_param, 'conformationindex', 'ebv_pzkop'),
                'gv': calculate_average(cows_param, 'conformationindex', 'ebv_gv'),
                'pdv': calculate_average(cows_param, 'conformationindex', 'ebv_pdv'),
                'vzcv': calculate_average(cows_param, 'conformationindex', 'ebv_vzcv'),
                'szcv': calculate_average(cows_param, 'conformationindex', 'ebv_szcv'),
                'csv': calculate_average(cows_param, 'conformationindex', 'ebv_csv'),
                'rps': calculate_average(cows_param, 'conformationindex', 'ebv_rps'),
                'rzs': calculate_average(cows_param, 'conformationindex', 'ebv_rzs'),
                'ds': calculate_average(cows_param, 'conformationindex', 'ebv_ds'),

                'milk': calculate_average(cows_param, 'milkproductionindex', 'ebv_milk'),
                'fkg': calculate_average(cows_param, 'milkproductionindex', 'ebv_fkg'),
                'fprc': calculate_average(cows_param, 'milkproductionindex', 'ebv_fprc'),
                'pkg': calculate_average(cows_param, 'milkproductionindex', 'ebv_pkg'),
                'pprc': calculate_average(cows_param, 'milkproductionindex', 'ebv_pprc'),

                'crh': calculate_average(cows_param, 'reproductionindex', 'ebv_crh'),
                'ctfi': calculate_average(cows_param, 'reproductionindex', 'ebv_ctfi'),
                'do': calculate_average(cows_param, 'reproductionindex', 'ebv_do'),

                'scs': calculate_average(cows_param, 'somaticcellindex', 'ebv_scs'),
            }

            forecasting_1 = set_predict(aggregated_data['forecasting_section_one'], forecasting, avg_dict)
            forecasting_2 = set_predict(aggregated_data['forecasting_section_two'], forecasting, avg_dict)
            forecasting_3 = set_predict(aggregated_data['forecasting_section_three'], forecasting, avg_dict)
            forecasting_4 = set_predict(aggregated_data['forecasting_section_four'], forecasting, avg_dict)
            farm.jsonfarmsdata.save()
            forecasting_1 = mapping_label(forecasting_1)
            forecasting_2 = mapping_label(forecasting_2)
            forecasting_3 = mapping_label(forecasting_3)
            forecasting_4 = mapping_label(forecasting_4)

            result_data = {
                'forecasting_data_one': forecasting_1,
                'forecasting_data_two': forecasting_2,
                'forecasting_data_thee': forecasting_3,
                'forecasting_data_four': forecasting_4,
            }

            return Response(result_data, status=status.HTTP_200_OK)
        except Farms.DoesNotExist:
            return Response({"error": "Проблема с данными"}, status=status.HTTP_400_BAD_REQUEST)
