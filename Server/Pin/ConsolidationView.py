import os
import re
from datetime import datetime
from django.conf import settings
from rest_framework import status
from transliterate import translit
from collections import defaultdict
from rest_framework.views import APIView
from rest_framework.response import Response
from ..models import Farms, PK, PKYoungAnimals, Parentage, PKBull, Report
from ..serializers import CowParameterForecastingSerializer, BullParameterForecastingSerializer

from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Image, Spacer, KeepTogether

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from .ParameterForecastingView import set_predict, calculate_weighted_average, calculate_weighted_average_with_bulls, \
    calculate_average, get_weighted_avg_bull, mapping_label

REPORT_DIR = os.path.join(settings.BASE_DIR, 'reports')
ICON_PATH = os.path.join(settings.BASE_DIR, 'image/dna.png')
FONT_PATH = os.path.join(settings.BASE_DIR, 'text/DejaVuSans.ttf')
FONT_BOLD_PATH = os.path.join(settings.BASE_DIR, 'text/DejaVuSans-Bold.ttf')


def perform_consolidation(cows, mod):
    if mod == 'cow':
        cows = PK.objects.filter(uniq_key__in=cows)
        for cow in cows:
            cow.consolidation = True
            cow.save()
    else:
        cows = PKYoungAnimals.objects.filter(uniq_key__in=cows)
        for cow in cows:
            cow.consolidation = True
            cow.save()


def get_ancestors_for_animals(animals, mode='bull', generations=3):
    if mode == 'cow' or mode == 'bull':
        ancestry = defaultdict(lambda: defaultdict(list))

        current_animals = set(animals)
        animals_with_few_relatives = []

        for generation in range(1, generations + 1):
            parentages = Parentage.objects.filter(uniq_key__in=current_animals)
            next_animals = set()
            for parentage in parentages:
                if parentage.ukeyo:
                    ancestry[parentage.uniq_key][generation].append(parentage.ukeyo)
                    next_animals.add(parentage.ukeyo)
                if parentage.ukeym:
                    ancestry[parentage.uniq_key][generation].append(parentage.ukeym)
                    next_animals.add(parentage.ukeym)
            current_animals = next_animals
    else:
        ancestry = defaultdict(lambda: defaultdict(list))

        parentages = PKYoungAnimals.objects.filter(uniq_key__in=animals)
        animals_with_few_relatives = []
        next_animals = set()

        for parentage in parentages:
            if parentage.f_regnomer:
                ancestry[parentage.uniq_key][1].append(parentage.f_regnomer)
                next_animals.add(parentage.f_regnomer)
            if parentage.m_regnomer:
                ancestry[parentage.uniq_key][1].append(parentage.m_regnomer)
                next_animals.add(parentage.m_regnomer)
        current_animals = next_animals

        for generation in range(2, generations + 1):
            parentages = Parentage.objects.filter(uniq_key__in=current_animals)
            next_animals = set()
            for parentage in parentages:
                if parentage.ukeyo:
                    ancestry[parentage.uniq_key][generation].append(parentage.ukeyo)
                    next_animals.add(parentage.ukeyo)
                if parentage.ukeym:
                    ancestry[parentage.uniq_key][generation].append(parentage.ukeym)
                    next_animals.add(parentage.ukeym)
            current_animals = next_animals

    def build_full_ancestry(animal, current_gen=1):
        """Рекурсивно добавляет предков в родословную."""
        if current_gen > generations or animal not in ancestry:
            return []

        parents = ancestry[animal].get(current_gen, [])

        grandparents = []
        for parent in parents:
            grandparents.extend(build_full_ancestry(parent, current_gen + 1))
        return parents + grandparents

    full_ancestry = {}

    for animal in animals:
        full_ancestry[animal] = {}
        parents = build_full_ancestry(animal)

        if len(parents) < 14:
            full_ancestry[animal][2] = parents
            full_ancestry[animal][3] = []
            full_ancestry[animal][4] = []

            animals_with_few_relatives.append({'animal': animal, 'relatives_count': len(parents)})
        else:
            full_ancestry[animal][2] = parents[0:2]
            full_ancestry[animal][3] = parents[2:4] + parents[8:10]
            full_ancestry[animal][4] = parents[4:8] + parents[10:14]

    return full_ancestry, animals_with_few_relatives


def check_inbreeding(bulls, cows, mode):
    """Функция проверяет на инбридинг для списка быков и коров"""
    results = {}

    ancestry_bull, not_full_bull = get_ancestors_for_animals(bulls)
    ancestry_cow, not_full_cow = get_ancestors_for_animals(cows, mode)

    for bull, bull_tree in ancestry_bull.items():
        for cow, cow_tree in ancestry_cow.items():
            for bull_level, bull_ancestors in bull_tree.items():
                for cow_level, cow_ancestors in cow_tree.items():
                    common_ancestors = set(bull_ancestors) & set(cow_ancestors)
                    if common_ancestors:
                        if bull not in results:
                            try:
                                bull_data = PKBull.objects.get(uniq_key=bull)
                                nomer = bull_data.nomer
                                klichka = bull_data.klichka
                            except PKBull.DoesNotExist:
                                nomer = ''
                                klichka = ''
                            except KeyError:
                                nomer = ''
                                klichka = ''
                            except Exception as e:
                                nomer = ''
                                klichka = ''
                                print(f"Произошла ошибка: {e}")
                            results[bull] = {
                                'bull': bull,
                                'nomer': nomer,  # Функция для получения номера
                                'klichka': klichka,  # Функция для получения клички
                                'inbreeding_cases': []
                            }
                        results[bull]['inbreeding_cases'].append({
                            'cow': cow,
                            'bull_level': bull_level,
                            'cow_level': cow_level,
                            'common_ancestors': list(common_ancestors)
                        })

    if len(results) < 1:
        return ['Нет инбредных животных']

    # not_full_cow.extend(not_full_bull)
    #
    # print(len(not_full_bull))
    # print(len(not_full_cow))
    #
    # print(list(results.values()))
    return list(results.values())


def sanitize_filename(name):
    """Преобразование названия в нижний регистр, замена пробелов на подчеркивания и транслитерация."""
    name = translit(name, 'ru', reversed=True)
    name = name.lower()  # Переводим в нижний регистр
    name = re.sub(r'\s+', '_', name)
    return name


def get_unique_filename(base_name, type_file, current_time=None):
    """Создание уникального имени файла с добавлением текущей даты и времени. Создание директории, если ее нет."""
    directory_path = os.path.join(REPORT_DIR, f'{base_name}')

    if not os.path.isdir(directory_path):
        os.mkdir(directory_path)
    if current_time is None:
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

    report_path = os.path.join(directory_path, f'{base_name}_{current_time}.{type_file}')

    return report_path, current_time


def create_pdf_report(cows, bulls, name_pdf, user_name=None):
    """Создание PDF отчета о закреплении коров за быками."""
    try:
        if not os.path.exists(REPORT_DIR):
            os.makedirs(REPORT_DIR)

        sanitized_name = sanitize_filename(name_pdf)
        pdf_path, current_time = get_unique_filename(sanitized_name, 'pdf')

        if user_name is not None:
            user_name = sanitize_filename(user_name)
            pdf_path = pdf_path[:-4] + '__' + user_name + pdf_path[-4:]

        pdfmetrics.registerFont(TTFont('DejaVu', FONT_PATH))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', FONT_BOLD_PATH))

        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=40, bottomMargin=30)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name='TitleStyle',
            fontName='DejaVu',
            fontSize=16,
            alignment=1,
            spaceAfter=12,
            leading=18
        )
        normal_style = ParagraphStyle(name='NormalStyle', fontName='DejaVu', fontSize=12, spaceAfter=4)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVu'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])

        try:
            logo = Image(ICON_PATH, width=30, height=30)
            logo.hAlign = 'LEFT'
        except Exception as e:
            print(f"Ошибка при загрузке иконки: {e}")
            logo = None

        title = Paragraph(
            f"Отчет о закреплении коров за быками в хозяйстве <font name='DejaVu-Bold' size=14>'{name_pdf}'</font>",
            title_style
        )

        header_table_data = [[logo, title]]
        header_table = Table(header_table_data, colWidths=[0.7 * inch, 5.3 * inch])
        header_table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'LEFT')]))

        current_date = datetime.now().strftime('%d %B %Y')
        current_date_russian = current_date.replace('January', 'Январь').replace('February', 'Февраль').replace(
            'March', 'Март').replace('April', 'Апрель').replace('May', 'Май').replace('June', 'Июнь').replace(
            'July', 'Июль').replace('August', 'Август').replace('September', 'Сентябрь').replace('October',
                                                                                                 'Октябрь').replace(
            'November', 'Ноябрь').replace('December', 'Декабрь')
        date_info = Paragraph(f"Дата: {current_date_russian}", normal_style)

        def draw_separator(canvas, doc):
            canvas.setStrokeColor(colors.black)
            canvas.setLineWidth(1)
            canvas.line(72, 710, 540, 710)

        def create_cow_table(data, title):
            cow_table = Table(data, colWidths=[0.5 * inch, 2 * inch], repeatRows=1, splitByRow=True)
            cow_table.setStyle(table_style)
            return cow_table

        if len(cows) > 33:
            first_cow_data = [['№', 'Коровы']] + [[str(idx), cow] for idx, cow in enumerate(cows[:33], start=1)]
            second_cow_data = [['№', 'Коровы']] + [[str(idx + 33), cow] for idx, cow in enumerate(cows[33:], start=1)]

            first_cow_table = create_cow_table(first_cow_data, 'Коровы')
            second_cow_table = create_cow_table(second_cow_data, 'Коровы')

            bull_data = [['№', 'Быки']]
            bull_data.extend([[str(idx), bull] for idx, bull in enumerate(bulls, start=1)])
            bull_table = create_cow_table(bull_data, 'Быки')

            animal_tables = Table([[first_cow_table, bull_table]], colWidths=[3 * inch, 3 * inch], splitByRow=True)
            animal_tables.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))

        else:
            cow_data = [['№', 'Коровы']] + [[str(idx), cow] for idx, cow in enumerate(cows, start=1)]
            cow_table = create_cow_table(cow_data, 'Коровы')

            bull_data = [['№', 'Быки']]
            bull_data.extend([[str(idx), bull] for idx, bull in enumerate(bulls, start=1)])
            bull_table = create_cow_table(bull_data, 'Быки')

            animal_tables = Table([[cow_table, bull_table]], colWidths=[3 * inch, 3 * inch], splitByRow=True)
            animal_tables.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))

        signature = Paragraph("Подпись: Белплемживобъединение", normal_style)
        if len(cows) > 33:

            content = [
                header_table,
                Spacer(1, 20),
                date_info,
                Spacer(1, 20),
                KeepTogether(animal_tables),
                Spacer(1, 20),
                second_cow_table,
                Spacer(1, 20),
                signature
            ]
        else:

            content = [
                header_table,
                Spacer(1, 20),
                date_info,
                Spacer(1, 20),
                KeepTogether(animal_tables),
                Spacer(1, 20),
                signature
            ]
        doc.build(content, onFirstPage=draw_separator)

        return pdf_path, current_time

    except Exception as e:
        print(f"Ошибка при создании PDF отчета: {e}")
        raise


def create_xlsx_report(cows, bulls, name_xlsx, mode, current_time, forecasting_1, forecasting_2, forecasting_3,
                       forecasting_4, user_name=None):
    try:
        if not os.path.exists(REPORT_DIR):
            os.makedirs(REPORT_DIR)

        sanitized_name = sanitize_filename(name_xlsx)
        xlsx_path, current_time = get_unique_filename(sanitized_name, 'xlsx', current_time)

        if user_name is not None:
            user_name = sanitize_filename(user_name)
            xlsx_path = xlsx_path[:-5] + '__' + user_name + xlsx_path[-5:]

        bulls_date = PKBull.objects.filter(uniq_key__in=bulls).values_list('nomer', 'uniq_key', 'klichka',
                                                                           'milkproductionindexbull__rm')

        if mode != 'young':
            cows = PK.objects.filter(uniq_key__in=cows).values_list('uniq_key', 'nomer', 'kodfer',
                                                                    'milkproductionindex__rm').order_by('kodfer')
        else:
            cows = PKYoungAnimals.objects.filter(uniq_key__in=cows).values_list('uniq_key', 'nomer', 'kodfer').order_by(
                'kodfer')

        wb = Workbook()
        ws = wb.active
        ws.title = "Закрепление"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Установка жирной границы для серой разделительной линии
        thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'),
                              bottom=Side(style='thick'))

        ws.merge_cells('A1:I2')
        ws['A1'] = name_xlsx
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['A1'].font = Font(bold=True)
        ws.merge_cells('A3:D3')
        ws['A3'] = 'Коровы'
        ws['A3'].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('F3:I3')
        ws['F3'] = 'Быки'
        ws['F3'].alignment = Alignment(horizontal="center", vertical="center")

        ws['A4'] = 'Индивидуальный номер'
        ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B4'] = 'Рабочий номер'
        ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C4'] = 'Код фермы'
        ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D4'] = 'RM'
        ws['D4'].alignment = Alignment(horizontal="center", vertical="center")

        ws['F4'] = 'Рабочий номер'
        ws['F4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G4'] = 'Индивидуальный номер'
        ws['G4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H4'] = 'Кличка'
        ws['H4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I4'] = 'RM'
        ws['I4'].alignment = Alignment(horizontal="center", vertical="center")

        for row in ws['A1:I2']:
            for cell in row:
                cell.border = thick_border
        for row in ws['F3:I3']:
            for cell in row:
                cell.border = thick_border
        for row in ws['F4:I4']:
            for cell in row:
                cell.border = thick_border
        for row in ws['A3:D4']:
            for cell in row:
                cell.border = thick_border

        # Установка ширины столбцов
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10

        ws.column_dimensions['E'].width = 15

        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 10

        row_start = 5

        for row_num, row_data in enumerate(cows, start=row_start):
            cell = ws.cell(row=row_num, column=1, value=row_data[0])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            cell = ws.cell(row=row_num, column=2, value=row_data[1])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            cell = ws.cell(row=row_num, column=3, value=row_data[2])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if mode != 'young':
                cell = ws.cell(row=row_num, column=4, value=row_data[3])
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_num, row_data in enumerate(bulls_date, start=row_start):
            cell = ws.cell(row=row_num, column=6, value=row_data[0])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            cell = ws.cell(row=row_num, column=7, value=row_data[1])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            cell = ws.cell(row=row_num, column=8, value=row_data[2])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            cell = ws.cell(row=row_num, column=9, value=row_data[3])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Прогноз закрепления ---
        forecast_headers = [
            ("Показатель", "param"),
            ("Средняя генетическая ценность коров", "avg"),
            ("Среднее генетическое превосходство быков над коровами", "bull_superiority"),
            ("Прогнозируемый эффект селекции потомства от скрещивания закрепленных коров и быков на поколение",
             "predict")
        ]

        def write_forecast_block(ws, forecast_data, start_col_index, start_row, title):
            from openpyxl.utils import get_column_letter

            # Заголовок таблицы
            end_col_index = start_col_index + len(forecast_headers) - 1
            start_col_letter = get_column_letter(start_col_index)
            end_col_letter = get_column_letter(end_col_index)

            ws.merge_cells(f"{start_col_letter}{start_row}:{end_col_letter}{start_row}")
            title_cell = ws[f"{start_col_letter}{start_row}"]
            title_cell.value = title
            title_cell.font = Font(bold=True, size=12)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Проставляем границу для всех ячеек объединённого заголовка
            for col in range(start_col_index, end_col_index + 1):
                col_letter = get_column_letter(col)
                cell = ws[f"{col_letter}{start_row}"]
                cell.border = thick_border

            # Заголовки
            for col_index, (header_name, _) in enumerate(forecast_headers):
                col_letter = get_column_letter(start_col_index + col_index)
                cell = ws[f"{col_letter}{start_row + 1}"]
                cell.value = header_name
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                cell.border = thick_border

                # Настройка ширины
                if col_index == 0:
                    ws.column_dimensions[col_letter].width = 40  # Широкая колонка для "Показатель"
                else:
                    ws.column_dimensions[col_letter].width = 22

            # Данные
            for row_offset, item in enumerate(forecast_data, start=2):
                for col_index, (_, key) in enumerate(forecast_headers):
                    col_letter = get_column_letter(start_col_index + col_index)
                    cell = ws[f"{col_letter}{start_row + row_offset}"]
                    value = item.get(key)
                    cell.value = round(value, 4) if isinstance(value, (int, float)) else value
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

        forecast_start_row = row_start
        write_forecast_block(ws, forecasting_1 + forecasting_2, 12, forecast_start_row, "Прогноз продуктивности")

        second_block_start_row = forecast_start_row + len(forecasting_1 + forecasting_2) + 5  # +5 для отступа
        write_forecast_block(ws, forecasting_3 + forecasting_4, 12, second_block_start_row, "Прогноз воспроизводства")

        wb.save(xlsx_path)

        return xlsx_path
    except Exception as e:
        print(f"Ошибка при создании xlsx отчета: {e}")
        raise


def forecasting_function(reports, aggregated_data):
    cows_param = []
    avg_values = []
    count_cows = 0

    for report in reports:
        cows = PK.objects.filter(uniq_key__in=report[0]).values_list('uniq_key', flat=True)

        if len(cows) == len(report[0]):
            queryset = PK.objects.filter(
                uniq_key__in=cows
            ).select_related(
                'milkproductionindex', 'conformationindex', 'reproductionindex',
                'somaticcellindex',
            ).order_by('id')

            number = queryset.count()
            count_cows += number
            serializer = CowParameterForecastingSerializer(queryset, many=True)
            cow_data = serializer.data

            queryset = PKBull.objects.filter(
                uniq_key__in=report[1]
            ).select_related(
                'milkproductionindexbull', 'conformationindexbull', 'reproductionindexbull',
                'somaticcellindexbull',
            ).order_by('id')

            serializer = BullParameterForecastingSerializer(queryset, many=True)
            bull_data = serializer.data

            average_conformation, avg_conf = calculate_weighted_average(bull_data, 'conformationindexbull',
                                                                        number)
            average_milk, avg_milk = calculate_weighted_average(bull_data, 'milkproductionindexbull', number)
            average_reproduction, avg_reprod = calculate_weighted_average(bull_data, 'reproductionindexbull',
                                                                          number)
            average_somaticcell, avg_somatic = calculate_weighted_average(bull_data, 'somaticcellindexbull',
                                                                          number)

            avg_values.append([avg_conf, avg_milk, avg_reprod, avg_somatic])
            cows = calculate_weighted_average_with_bulls(cow_data,
                                                         [average_conformation, average_milk,
                                                          average_reproduction,
                                                          average_somaticcell],
                                                         ['conformationindex', 'milkproductionindex',
                                                          'reproductionindex', 'somaticcellindex'])

            cows_param.extend(cows)
        else:
            pass
    avg_dict = get_weighted_avg_bull(avg_values, count_cows)

    forecasting = {
        'tip': calculate_average(cows_param, 'conformationindex', 'ebv_tip'),
        'kt': calculate_average(cows_param, 'conformationindex', 'ebv_kt'),
        'rost': calculate_average(cows_param, 'conformationindex', 'ebv_rost'),
        'gt': calculate_average(cows_param, 'conformationindex', 'ebv_gt'),
        'pz': calculate_average(cows_param, 'conformationindex', 'ebv_pz'),
        'shz': calculate_average(cows_param, 'conformationindex', 'ebv_shz'),
        'pzkb': calculate_average(cows_param, 'conformationindex', 'ebv_pzkb'),
        'pzkz': calculate_average(cows_param, 'conformationindex', 'ebv_pzkz'),
        'sust': calculate_average(cows_param, 'conformationindex', 'ebv_sust'),
        'pzkop': calculate_average(cows_param, 'conformationindex', 'ebv_pzkop'),
        'gv': calculate_average(cows_param, 'conformationindex', 'ebv_gv'),
        'pdv': calculate_average(cows_param, 'conformationindex', 'ebv_pdv'),
        'vzcv': calculate_average(cows_param, 'conformationindex', 'ebv_vzcv'),
        'szcv': calculate_average(cows_param, 'conformationindex', 'ebv_szcv'),
        'csv': calculate_average(cows_param, 'conformationindex', 'ebv_csv'),
        'rps': calculate_average(cows_param, 'conformationindex', 'ebv_rps'),
        'rzs': calculate_average(cows_param, 'conformationindex', 'ebv_rzs'),
        'ds': calculate_average(cows_param, 'conformationindex', 'ebv_ds'),

        'milk': calculate_average(cows_param, 'milkproductionindex', 'ebv_milk'),
        'fkg': calculate_average(cows_param, 'milkproductionindex', 'ebv_fkg'),
        'fprc': calculate_average(cows_param, 'milkproductionindex', 'ebv_fprc'),
        'pkg': calculate_average(cows_param, 'milkproductionindex', 'ebv_pkg'),
        'pprc': calculate_average(cows_param, 'milkproductionindex', 'ebv_pprc'),

        'crh': calculate_average(cows_param, 'reproductionindex', 'ebv_crh'),
        'ctfi': calculate_average(cows_param, 'reproductionindex', 'ebv_ctfi'),
        'do': calculate_average(cows_param, 'reproductionindex', 'ebv_do'),

        'scs': calculate_average(cows_param, 'somaticcellindex', 'ebv_scs'),
    }

    forecasting_1 = set_predict(aggregated_data['forecasting_section_one'], forecasting, avg_dict)
    forecasting_2 = set_predict(aggregated_data['forecasting_section_two'], forecasting, avg_dict)
    forecasting_3 = set_predict(aggregated_data['forecasting_section_three'], forecasting, avg_dict)
    forecasting_4 = set_predict(aggregated_data['forecasting_section_four'], forecasting, avg_dict)

    forecasting_1 = mapping_label(forecasting_1)
    forecasting_2 = mapping_label(forecasting_2)
    forecasting_3 = mapping_label(forecasting_3)
    forecasting_4 = mapping_label(forecasting_4)

    return forecasting_1, forecasting_2, forecasting_3, forecasting_4


class ConsolidationView(APIView):

    def post(self, request):
        data = request.data
        user_name = data['name']
        user = request.user  # Получаем текущего пользователя из токена
        cows = data['cows']
        bulls = data['bulls']
        mode = data['mode']

        try:
            name_pdf = Farms.objects.get(korg=self.request.headers.get('Kodrn')).norg
            farm = Farms.objects.get(korg=self.request.headers.get('Kodrn'))
            aggregated_data = farm.jsonfarmsdata.aggregated_data['aggregated_data']
            mod = self.request.headers.get('Mode')
        except Farms.DoesNotExist:
            return Response({"error": "Ферма не найдена."}, status=status.HTTP_404_NOT_FOUND)
        try:
            if mod == 'standard':
                inbreeding_results = check_inbreeding(bulls, cows, mode)
                if len(inbreeding_results) == 1 and inbreeding_results[0] == 'Нет инбредных животных':
                    return Response({
                        "inbreeding_check": True,
                    }, status=status.HTTP_200_OK)
                else:
                    response_data = []
                    for element in inbreeding_results:
                        response_data.append({
                            "bull": element['bull'],
                            "inbred_cows_count": len(element['inbreeding_cases']),
                        })
                    return Response({
                        "inbreeding_check": False,
                        "inbred_animals": inbreeding_results,
                        "count_unique_cows": response_data
                    }, status=status.HTTP_200_OK)
            elif mod == "With":
                perform_consolidation(cows, mode)
                if user_name:
                    pdf_file_path, current_time = create_pdf_report(cows, bulls, name_pdf, user_name)
                    forecasting_1, forecasting_2, forecasting_3, forecasting_4 = forecasting_function([[cows, bulls]],
                                                                                                      aggregated_data)
                    create_xlsx_report(cows, bulls, name_pdf, mode, current_time, forecasting_1, forecasting_2,
                                       forecasting_3, forecasting_4, user_name)
                    path = pdf_file_path.replace('.pdf', '')
                    path = os.path.basename(path)
                    Report.objects.create(
                        title=f"{name_pdf} ({user_name})",
                        user=user,
                        path=path,
                    )

                else:
                    pdf_file_path, current_time = create_pdf_report(cows, bulls, name_pdf)
                    forecasting_1, forecasting_2, forecasting_3, forecasting_4 = forecasting_function([[cows, bulls]],
                                                                                                      aggregated_data)
                    create_xlsx_report(cows, bulls, name_pdf, mode, current_time, forecasting_1, forecasting_2,
                                       forecasting_3, forecasting_4)

                    path = pdf_file_path.replace('.pdf', '')
                    path = os.path.basename(path)
                    Report.objects.create(
                        title=f"{name_pdf} ({user_name})",
                        user=user,
                        path=path,
                    )

                return Response({
                    "inbreeding_check": True,
                    "pdf_filename": os.path.basename(pdf_file_path)
                }, status=status.HTTP_200_OK)
            elif mod == "Without":
                without = data['inbred']
                cows_to_remove = [
                    case['cow']
                    for entry in without
                    for case in entry.get('inbreeding_cases', [])
                ]
                filtered_cows = [cow for cow in cows if cow not in cows_to_remove]
                perform_consolidation(filtered_cows, mode)
                if user_name:
                    pdf_file_path, current_time = create_pdf_report(filtered_cows, bulls, name_pdf, user_name)
                    forecasting_1, forecasting_2, forecasting_3, forecasting_4 = forecasting_function([[cows, bulls]],
                                                                                                      aggregated_data)
                    create_xlsx_report(filtered_cows, bulls, name_pdf, mode, current_time, forecasting_1, forecasting_2,
                                       forecasting_3, forecasting_4, user_name)

                    path = pdf_file_path.replace('.pdf', '')
                    path = os.path.basename(path)
                    Report.objects.create(
                        title=f"{name_pdf} ({user_name})",
                        user=user,
                        path=path,
                    )

                else:
                    pdf_file_path, current_time = create_pdf_report(filtered_cows, bulls, name_pdf)
                    forecasting_1, forecasting_2, forecasting_3, forecasting_4 = forecasting_function([[cows, bulls]],
                                                                                                      aggregated_data)
                    create_xlsx_report(filtered_cows, bulls, name_pdf, mode, current_time, forecasting_1, forecasting_2,
                                       forecasting_3, forecasting_4)

                    path = pdf_file_path.replace('.pdf', '')
                    path = os.path.basename(path)
                    Report.objects.create(
                        title=f"{name_pdf} ({user_name})",
                        user=user,
                        path=path,
                    )

                return Response({
                    "inbreeding_check": True,
                    "pdf_filename": os.path.basename(pdf_file_path)
                }, status=status.HTTP_200_OK)
            elif mod == 'standard_confirm':
                inbreeding_results = check_inbreeding(bulls, cows, mode)
                if len(inbreeding_results) == 1 and inbreeding_results[0] == 'Нет инбредных животных':
                    perform_consolidation(cows, mode)
                    if user_name:
                        pdf_file_path, current_time, = create_pdf_report(cows, bulls, name_pdf, user_name)
                        forecasting_1, forecasting_2, forecasting_3, forecasting_4 = forecasting_function(
                            [[cows, bulls]], aggregated_data)
                        create_xlsx_report(cows, bulls, name_pdf, mode, current_time, forecasting_1, forecasting_2,
                                           forecasting_3, forecasting_4, user_name)
                        path = pdf_file_path.replace('.pdf', '')
                        path = os.path.basename(path)
                        Report.objects.create(
                            title=f"{name_pdf} ({user_name})",
                            user=user,
                            path=path,
                        )

                    else:
                        pdf_file_path, current_time, = create_pdf_report(cows, bulls, name_pdf)
                        forecasting_1, forecasting_2, forecasting_3, forecasting_4 = forecasting_function(
                            [[cows, bulls]], aggregated_data)
                        create_xlsx_report(cows, bulls, name_pdf, mode, current_time, forecasting_1, forecasting_2,
                                           forecasting_3, forecasting_4)

                        path = pdf_file_path.replace('.pdf', '')
                        path = os.path.basename(path)
                        Report.objects.create(
                            title=f"{name_pdf} ({user_name})",
                            user=user,
                            path=path,
                        )

                    return Response({
                        "inbreeding_check": True,
                        "pdf_filename": os.path.basename(pdf_file_path)  # Имя PDF файла для фронтенда
                    }, status=status.HTTP_200_OK)
                else:
                    return Response({
                        "inbreeding_check": False,
                        "inbred_animals": inbreeding_results
                    }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
