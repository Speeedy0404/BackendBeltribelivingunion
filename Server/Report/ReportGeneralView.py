import os

from copy import copy
from django.conf import settings

from ..models import Report
from rest_framework import status
from django.http import FileResponse
from rest_framework.views import APIView
from rest_framework.response import Response

from openpyxl import Workbook, load_workbook
from io import BytesIO

REPORT_DIR = os.path.join(settings.BASE_DIR, 'reports')


class ReportGeneralView(APIView):
    def get(self, request, farm):
        if not farm:
            return Response({"error": "Хозяйство не найдено."}, status=status.HTTP_404_NOT_FOUND)

        reports = Report.objects.filter(title__icontains=farm)
        if not reports.exists():
            return Response({"error": "Нет отчетов для данного хозяйства."}, status=status.HTTP_404_NOT_FOUND)

        combined_wb = Workbook()
        combined_wb.remove(combined_wb.active)

        for report in reports:
            print(report.path)

            if '__' in report.path:
                directory_path = report.path.split('__')[0]
            else:
                directory_path = report.path.split('.')[0]
            directory_path = directory_path[:-16]

            report_path = os.path.join(REPORT_DIR, directory_path, report.path + '.xlsx')

            if not os.path.exists(report_path):
                continue

            wb = load_workbook(report_path)

            try:
                for sheet in wb.worksheets:

                    new_sheet_title = f"{report.title[:28]}".strip()
                    new_sheet = combined_wb.create_sheet(title=new_sheet_title)

                    for row in sheet.iter_rows():
                        for cell in row:
                            new_cell = new_sheet.cell(
                                row=cell.row,
                                column=cell.column,
                                value=cell.value
                            )
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = cell.number_format
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)

                    for col_idx, col_dim in sheet.column_dimensions.items():
                        new_sheet.column_dimensions[col_idx].width = col_dim.width
            except Exception as e:
                print(f"Ошибка при обработке файла {report.path}: {e}")
                continue

        output = BytesIO()
        combined_wb.save(output)
        output.seek(0)

        filename = f"{farm}_общий_отчет.xlsx"
        return FileResponse(output, as_attachment=True, filename=filename)
