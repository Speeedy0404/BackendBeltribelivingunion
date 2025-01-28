import os
from django.conf import settings
from openpyxl import load_workbook
from rest_framework import status
from django.http import FileResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from ..models import Report, PK, PKYoungAnimals

REPORT_DIR = os.path.join(settings.BASE_DIR, 'reports')


class ReportView(APIView):

    def get(self, request, filename):
        """Отправка PDF файла для скачивания по его имени."""
        if '__' in filename:
            directory_path = filename.split('__')[0]
        else:
            directory_path = filename.split('.')[0]

        directory_path = directory_path[:-16]

        report_path = os.path.join(REPORT_DIR, directory_path)
        report_path = os.path.join(report_path, filename)
        if os.path.exists(report_path):
            return FileResponse(open(report_path, 'rb'), as_attachment=True, filename=filename)
        else:
            return Response({"error": "Отчет не найден."}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, filename):

        cow_numbers = []
        search_name = filename.replace('.pdf', '')
        user = request.user

        try:
            report = Report.objects.get(path=search_name)
        except Report.DoesNotExist:
            return Response({"success": False, "error": "Отчет не найден."}, status=status.HTTP_200_OK)

        try:

            if user == report.user:
                if '__' in filename:
                    directory_path = filename.split('__')[0]
                else:
                    directory_path = filename.split('.')[0]
                directory_path = directory_path[:-16]
                report_path = os.path.join(REPORT_DIR, directory_path, filename)
                xlsx_path = report_path.replace('pdf', 'xlsx')

                if os.path.exists(report_path):
                    os.remove(report_path)
                try:
                    if os.path.exists(xlsx_path):

                        workbook = load_workbook(xlsx_path)
                        sheet = workbook.active

                        for row in sheet.iter_rows(min_row=5, max_col=1, values_only=True):
                            if row[0]:
                                cow_numbers.append(row[0])


                except Exception as e:
                    print(f"Ошибка при удалении файла: {e}")

                cows = PK.objects.filter(uniq_key__in=cow_numbers).values_list('uniq_key')

                if len(cows) == len(cow_numbers):
                    PK.objects.filter(uniq_key__in=cow_numbers).update(consolidation=False)
                else:
                    PKYoungAnimals.objects.filter(uniq_key__in=cow_numbers).update(consolidation=False)

                if os.path.exists(xlsx_path):
                    os.remove(xlsx_path)

                report.delete()

                return Response({"success": True, "message": "Отчет успешно удален."}, status=status.HTTP_200_OK)
            else:
                return Response({"success": False, "message": "Нет доступа для удаления отчета."},
                                status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"success": False, "error": f"Ошибка при удалении отчета: {str(e)}"},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)
